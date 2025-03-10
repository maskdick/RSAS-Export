import os
import glob
import zipfile
from lxml import etree
import json
import re
import openpyxl

def to_string(ip):
    return etree.tostring(ip, encoding='utf-8', method='html').decode('utf-8')

def process_html_content(html_str, ws_vul, ws_status, start_row_vul, start_row_status):
    result = etree.HTML(html_str)

    j_son = result.xpath('//script[1]')
    for i in j_son:
        a = to_string(i)

    pattern = r'<script>window\.data\s*=\s*({.*?});</script>'
    match = re.search(pattern, a, re.DOTALL)

    if match:
        json_str = match.group(1)
        try:
            data = json.loads(json_str)
            for item in data['categories'][1]['children'][0]['data']['vul_info']['vul_items']:
                service = item['service']
                protocol = item['protocol']
                for vul in item['vuls']:
                    vul_level = vul['vul_level']
                    if vul_level == "low":
                        vul['vul_level'] = "低"
                    elif vul_level == "middle":
                        vul['vul_level'] = "中"
                    elif vul_level == "high":
                        vul['vul_level'] = "高"
                    vul_msg = vul['vul_msg']
                    host_ip = vul_msg['host_ip']
                    port = vul['port']
                    vul_level = vul['vul_level']
                    cve = vul_msg['cve_id']
                    i18n_name = vul_msg['i18n_name']
                    i18n_description = "\n".join(vul_msg['i18n_description'])
                    i18n_solution = "\n".join(vul_msg['i18n_solution'])

                    ws_vul.append([start_row_vul, host_ip, port, protocol, service, i18n_name, vul_level, i18n_solution, i18n_description, cve])
                    start_row_vul += 1

            for status in data['categories'][4]['data']['other_info_data']:
                if status['info_name'] == '远程端口信息':
                    content = status.get('content', [])
                    for s in content:
                        if len(s) >= 4:
                            sport = s[0]
                            sprotocol = s[1]
                            sservice = s[2]
                            sstatus = s[3]
                            shost_ip = data['categories'][0]['data']['target']

                            # 检查并拆分端口范围
                            if '-' in sport:
                                start_port, end_port = map(int, sport.split('-'))
                                for port in range(start_port, end_port + 1):
                                    ws_status.append([shost_ip, int(port), sprotocol, sservice, sstatus])
                                    start_row_status += 1
                            else:
                                ws_status.append([shost_ip, int(sport), sprotocol, sservice, sstatus])
                                start_row_status += 1

        except json.JSONDecodeError as e:
            print("JSON解析错误:", e)
    else:
        print("未找到匹配的JSON内容")
    return start_row_vul, start_row_status

def process_zip_file(zip_path, ws_vul, ws_status, start_row_vul, start_row_status):
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        file_list = zip_ref.namelist()
        html_files = [f for f in file_list if f.startswith('host/') and f.endswith('.html')]

        for html_file in html_files:
            with zip_ref.open(html_file) as f:
                html_str = f.read().decode('utf-8')
                start_row_vul, start_row_status = process_html_content(html_str, ws_vul, ws_status, start_row_vul, start_row_status)

    return start_row_vul, start_row_status

def main():
    directory = "./"  # 请将此处替换为实际目录路径
    zip_files = glob.glob(os.path.join(directory, '*.zip'))

    print("请选择输出模式：")
    print("1. 合并所有报告到一个Excel文件")
    print("2. 每个ZIP文件生成单独的Excel报告")
    choice = input("请输入1或2，直接回车默认选择1：")

    if choice == '2':
        for zip_path in zip_files:
            process_individual_zip(zip_path, directory)
    else:
        process_combined_zips(zip_files, directory)

def process_combined_zips(zip_files, directory):
    wb_vul = openpyxl.Workbook()
    ws_vul = wb_vul.active
    ws_vul.title = "系统漏洞"

    wb_status = openpyxl.Workbook()
    ws_status = wb_status.active
    ws_status.title = "端口状态"

    header_vul = ["序号", "IP地址", "端口", "协议", "服务", "漏洞名称", "风险等级", "整改建议", "漏洞描述", "漏洞CVE编号"]
    ws_vul.append(header_vul)

    header_status = ["IP", "端口", "协议", "服务", "开放状态"]
    ws_status.append(header_status)

    start_row_vul = 1
    start_row_status = 1

    for zip_path in zip_files:
        start_row_vul, start_row_status = process_zip_file(zip_path, ws_vul, ws_status, start_row_vul, start_row_status)

    output_dir = os.path.join(directory, "combined_reports")
    os.makedirs(output_dir, exist_ok=True)

    excel_path_vul = os.path.join(output_dir, "漏洞报告.xlsx")
    wb_vul.save(excel_path_vul)
    print(f"保存Excel文件: {excel_path_vul}")

    excel_path_status = os.path.join(output_dir, "开放端口.xlsx")
    wb_status.save(excel_path_status)
    print(f"保存Excel文件: {excel_path_status}")

def process_individual_zip(zip_path, directory):
    wb_vul = openpyxl.Workbook()
    ws_vul = wb_vul.active
    ws_vul.title = "系统漏洞"

    wb_status = openpyxl.Workbook()
    ws_status = wb_status.active
    ws_status.title = "端口状态"

    header_vul = ["序号", "IP地址", "端口", "协议", "服务", "漏洞名称", "风险等级", "整改建议", "漏洞描述", "漏洞CVE编号"]
    ws_vul.append(header_vul)

    header_status = ["IP", "端口", "协议", "服务", "开放状态"]
    ws_status.append(header_status)

    start_row_vul = 1
    start_row_status = 1

    start_row_vul, start_row_status = process_zip_file(zip_path, ws_vul, ws_status, start_row_vul, start_row_status)

    zip_name = os.path.splitext(os.path.basename(zip_path))[0]
    output_dir = os.path.join(directory, zip_name)
    os.makedirs(output_dir, exist_ok=True)

    excel_path_vul = os.path.join(output_dir, "漏洞报告.xlsx")
    wb_vul.save(excel_path_vul)
    print(f"保存Excel文件: {excel_path_vul}")

    excel_path_status = os.path.join(output_dir, "开放端口.xlsx")
    wb_status.save(excel_path_status)
    print(f"保存Excel文件: {excel_path_status}")

if __name__ == "__main__":
    main()
